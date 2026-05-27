from __future__ import annotations

from pathlib import Path
from typing import Optional, Dict
import re
import csv
from copy import copy
import math

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from src.engine.phase_sums import update_single_phase_sumifs
from src.engine.busbars import rebuild_busbar_formulas


SHEET_NAME = "расчет"

COL_SECTION = "A"
COL_KPD = "B"
COL_TYPE = "C"
COL_NAME = "D"
COL_U = "E"
COL_PHASES = "F"
COL_P = "G"
COL_I = "H"
COL_KI = "I"
COL_COS = "J"
COL_TG = "K"
COL_L = "L"
COL_M = "M"
COL_N = "N"

# Итоговые колонки секций/ВРУ
COL_SEC_P_ACTIVE = "Q"
COL_SEC_Q_REACTIVE = "R"
COL_SEC_S_APPARENT = "S"
COL_SEC_I_TOTAL = "T"

# В строке УКРМ обычно используется R (кВАр)
COL_UKRM_KVAR = "R"

DATA_START_ROW = 5
LOADS_START_ROW = 6

RESERVE_NAME_RE = re.compile(r"^\s*резерв\s*$", re.IGNORECASE)
SERVICE_ROW_RE = re.compile(r"(секция\s+шин|итого\s+вру|регулируемое\s+устройство|компенсац)", re.IGNORECASE)

PUMP_FAMILIES: Dict[str, re.Pattern] = {
    "pump:podpitka": re.compile(r"насос.*подпиточ", re.IGNORECASE),
    "pump:ishod": re.compile(r"насос.*исходн", re.IGNORECASE),
    "pump:setev": re.compile(r"насос.*сетев", re.IGNORECASE),
    "pump:kotel_circ": re.compile(r"насос.*циркуляц.*котл", re.IGNORECASE),
    "pump:gvs_circ": re.compile(r"насос.*циркуляц.*гвс", re.IGNORECASE),
    "pump:any": re.compile(r"\bнасос\b", re.IGNORECASE),
}

TAG_ONLY_RE = re.compile(r"^\s*К\d{1,3}(?:\.\d{1,2})?\s*$", re.IGNORECASE)


def _clean_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").replace("\u00A0", " ")).strip()


def _norm(s: str) -> str:
    s = (s or "").replace("K", "К").replace("k", "К")
    s = re.sub(r"\s+", "", s)
    return s

def _norm_text(s: str | None) -> str:
    s = str(s or "").strip()
    s = s.replace("K", "К").replace("k", "К")
    s = re.sub(r"\s+", "", s)
    return s.upper()


def _parse_int_tail(tag: str) -> int | None:
    m = re.search(r"(\d+)$", str(tag or ""))
    return int(m.group(1)) if m else None


def _excel_tag_from_item(it: dict) -> str | None:
    """
    Переводит внутренний tag/item в обозначение строки Excel.
    Основная цель — находить уже существующие строки шаблона,
    а не плодить APPEND.
    """
    
    canonical = str(it.get("excel_tag_canonical") or "").strip()
    if canonical:
        return canonical
    
    tag = str(it.get("tag") or "").strip()
    ep_kind = str(it.get("ep_kind") or "").strip().lower()
    display_name = str(it.get("display_name") or "").strip().lower()
    duty = str(it.get("duty") or "").strip().lower()

    if not tag:
        return None

    # уже excel-friendly
    if re.fullmatch(r"(ШУК|ЩУТ|ЩУГ|ШСС|ХВО|ПЭСПЗ|ВН|ШК\d+|ЕК\d+|ГР\.\d+|Г\d+|А\d+|М\d+)", _norm_text(tag)):
        return tag

    # Горелки: ГГ.1 -> Г1
    if tag.upper().startswith("ГГ"):
        n = _parse_int_tail(tag)
        if n is not None:
            return f"Г{n}"

    # Шкаф управления котлом оставляем как есть: ШК1, ШК2...
    if tag.upper().startswith("ШК"):
        n = _parse_int_tail(tag)
        if n is not None:
            return f"ШК{n}"

    # Освещение: Гр.1...
    if tag.upper().startswith("ГР"):
        n = _parse_int_tail(tag)
        if n is not None:
            return f"Гр.{n}"

    # Электрообогрев: EK1 -> ЕК1
    if tag.upper().startswith(("EK", "ЕК")):
        n = _parse_int_tail(tag)
        if n is not None:
            return f"ЕК{n}"

    # HVAC агрегаты: А1...
    if ep_kind == "hvac":
        # агрегаты отопительные А1-А4 оставляем А1...
        if re.match(r"^[АA]\d+$", tag, flags=re.IGNORECASE):
            n = _parse_int_tail(tag)
            if n is not None:
                return f"А{n}"

        # вентиляторы/прочее ОВ пока оставляем по исходному tag
        # чтобы не сломать уже существующие UPDATE для П1-П4, В1
        if re.match(r"^[ПPВV]\d+$", tag, flags=re.IGNORECASE):
            return tag.replace("P", "П").replace("V", "В")

    # Насосы: К4/К5/К6... -> М1/М2/... по display_name/порядку
    # Здесь не пытаемся угадать через исходный tag. Для насосов основной fallback уже есть:
    # _find_row_by_model / _find_row_by_family.
    # Поэтому возвращаем None, чтобы не сделать ложное сопоставление.
    if ep_kind == "pump":
        return None

    # Пользовательские шкафы и отдельные позиции
    if tag in {"ШУК", "ЩУТ", "ЩУГ", "ШСС", "ХВО", "ПЭСПЗ", "ВН"}:
        return tag

    return None

def _full_tag(tag: str) -> str:
    # Полный тег (с .1/.2), для дедупа
    return _norm(tag)

def _canon_tag_for_keep(tag: str) -> str:
    """
    Канонизация тега для сравнения с keep_tags:
    - К7.2 -> К7 (насосные под-теги в скобках)
    - остальное оставляем как есть
    """
    t = _norm(tag)
    t = t.replace("ЩУТ", "ШУТ").replace("ЩУГ", "ШУГ")
    m = re.match(r"^(к\d{1,3})(?:\.\d{1,2})?$", t, re.IGNORECASE)
    if m:
        return m.group(1)  # К7.2 -> К7
    return t

def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _row_is_service(name: str) -> bool:
    return bool(name) and bool(SERVICE_ROW_RE.search(str(name)))

def _find_rows_by_excel_tag(ws, excel_tag: str, start_row: int, end_row: int) -> list[int]:
    """
    Ищет строку по колонке типа/обозначения (Г1, М1, ШК1, Гр.1 ...).
    """
    target = _norm_text(excel_tag)
    rows = []

    for r in range(start_row, end_row + 1):
        v_type = ws[f"{COL_TYPE}{r}"].value
        v_name = ws[f"{COL_NAME}{r}"].value

        if _norm_text(v_type) == target or _norm_text(v_name) == target:
            rows.append(r)

    return rows

def _guess_family_from_model(model: Optional[str]) -> str | None:
    if not model:
        return None
    m = _clean_ws(model).lower()

    # Подпиточный (К9)
    if "cdm3-7" in m or ("cdm3" in m and "7" in m):
        return "pump:podpitka"

    # Исходной воды (К8)
    if "cdm42-4" in m or ("cdm42" in m and "4" in m):
        return "pump:ishod"

    # Сетевой (К5)
    if "td200" in m:
        return "pump:setev"

    # ГВС внешний/внутренний (К6/К7)
    if "chl12-40" in m or "llts80-5-1.5" in m or "llts80" in m:
        return "pump:gvs_circ"

    # Циркуляция котла (К4)
    if "nis150" in m or ("125" in m and "200" in m):
        return "pump:kotel_circ"

    return None


def _family_key_from_name(name: str) -> str | None:
    s = _clean_ws(name).lower()
    if not s:
        return None
    for k in ("pump:podpitka", "pump:ishod", "pump:setev", "pump:kotel_circ", "pump:gvs_circ"):
        if PUMP_FAMILIES[k].search(s):
            return k
    if PUMP_FAMILIES["pump:any"].search(s):
        return "pump:any"
    return None


def _family_key_from_item(display_name: str, model: Optional[str]) -> str | None:
    fam = _family_key_from_name(display_name)
    if fam:
        return fam

    # если display_name == "К9" (голый тег) — пробуем через модель
    if display_name and TAG_ONLY_RE.match(display_name):
        fam2 = _guess_family_from_model(model)
        if fam2:
            return fam2

    return None


def _copy_row_style(ws, src_row: int, dst_row: int, *, min_col: int = 1, max_col: int = 60):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(min_col, max_col + 1):
        src: Cell = ws.cell(row=src_row, column=c)
        dst: Cell = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def find_busbar_block_start(ws, *, start_row: int = 1, end_row: int = 5000) -> int | None:
    pattern = re.compile(r"(?i)секции\s+шин|секция\s+шин")
    for r in range(start_row, end_row + 1):
        v = ws[f"{COL_NAME}{r}"].value
        if v and pattern.search(str(v)):
            return r
    return None


def find_reserve_row(ws, *, start_row: int, end_row: int) -> int | None:
    for r in range(start_row, end_row + 1):
        v = ws[f"{COL_NAME}{r}"].value
        if v and RESERVE_NAME_RE.match(str(v)):
            return r
    return None


def _audit_write(audit_writer: Optional[csv.writer], tag: str, addr: str, old, new, source: str):
    if audit_writer is None:
        return
    audit_writer.writerow([tag, addr, old, new, source])


def _set_cell(ws, addr: str, value, *, overwrite: bool, log: list[str], tag: str,
              audit_writer: Optional[csv.writer], source: str):
    if value is None:
        return

    cell = ws[addr]
    old = cell.value

    if _is_formula(old):
        log.append(f"[SKIP_FORMULA] {tag} {addr}: keep_formula")
        return

    if old in (None, ""):
        cell.value = value
        log.append(f"[WRITE] {tag} {addr}: {value}")
        _audit_write(audit_writer, tag, addr, old, value, source)
        return

    if overwrite and old != value:
        cell.value = value
        log.append(f"[OVERWRITE] {tag} {addr}: {old} -> {value}")
        _audit_write(audit_writer, tag, addr, old, value, source)
    elif old != value:
        log.append(f"[SKIP] {tag} {addr}: template={old} new={value} source={source}")


def apply_load_row_formulas(ws, r: int, log: list[str]):
    if ws[f"{COL_KI}{r}"].value in (None, ""):
        ws[f"{COL_KI}{r}"].value = 1

    # Ток (A) — как у тебя принято: для 220 и 380
    ws[f"{COL_I}{r}"].value = (
        f"=IF(E{r}<=230,"
        f"G{r}/((E{r}/1000)*J{r}*(B{r}/100)),"
        f"G{r}/(1.732*(E{r}/1000)*J{r}*(B{r}/100)))"
    )

    # tgφ
    ws[f"{COL_TG}{r}"].value = f"=ROUND(SQRT(1-(J{r})^2)/J{r},2)"

    # служебные расчёты (если используются в шаблоне)
    ws[f"{COL_L}{r}"].value = f"=I{r}*G{r}"
    ws[f"{COL_M}{r}"].value = f"=I{r}*G{r}*K{r}"
    ws[f"{COL_N}{r}"].value = f"=1*G{r}^2"


def _parse_section(v) -> int | None:
    try:
        x = int(str(v).strip())
        return x if x in (11, 12) else None
    except Exception:
        return None


def _parse_phase(v) -> str | None:
    s = str(v).strip().upper()
    return s if s in ("L1", "L2", "L3") else None


def _choose_section(section_p_sums: dict[int, float]) -> int:
    return 11 if section_p_sums.get(11, 0.0) <= section_p_sums.get(12, 0.0) else 12


def _choose_phase(phase_p_sums: dict[str, float]) -> str:
    return min(("L1", "L2", "L3"), key=lambda ph: phase_p_sums.get(ph, 0.0))

def _find_rows_by_tag_any(ws, tag: str, *, start_row: int, end_row: int) -> list[int]:
    t = _norm(tag)

    # Для насосов К5 -> допускаем К5.1, К5.2 и т.д.
    if re.match(r"^К\d{1,3}$", t):
        pattern = re.compile(rf"(?<!\w){re.escape(t)}(?:\.\d+)?(?!\w)", re.IGNORECASE)
    else:
        # Для остальных тегов запрещаем ложные матчи в подстроках вида В1.1, П1.2 и т.п.
        pattern = re.compile(rf"(?<!\w){re.escape(t)}(?!\.\d)(?!\w)", re.IGNORECASE)

    rows = set()
    for r in range(start_row, end_row + 1):
        for col in (COL_NAME, COL_TYPE):
            v = ws[f"{col}{r}"].value
            if not v:
                continue

            sv = _norm(str(v))

            # exact match по колонке типа оставляем приоритетным
            if col == COL_TYPE and sv == t:
                rows.add(r)
                continue

            if pattern.search(sv):
                rows.add(r)

    return sorted(rows)

# def _find_rows_by_tag_any(ws, tag: str, *, start_row: int, end_row: int) -> list[int]:
#     t = _norm(tag)
#     if re.match(r"^К\d{1,3}$", t):
#         pattern = re.compile(rf"(?<!\w){re.escape(t)}(?:\.\d+)?(?!\w)")
#     else:
#         pattern = re.compile(rf"(?<!\w){re.escape(t)}(?!\w)")

#     rows = set()
#     for r in range(start_row, end_row + 1):
#         for col in (COL_NAME, COL_TYPE):
#             v = ws[f"{col}{r}"].value
#             if not v:
#                 continue

#             sv = _norm(str(v))

#             if col == COL_TYPE and sv == t:
#                 rows.add(r)
#                 continue

#             # остальное — regex (по имени и другим случаям)
#             if pattern.search(sv):
#                 rows.add(r)
#     return sorted(rows)

def _find_pump_group_rows(ws, pump_tag: str, *, start_row: int, end_row: int) -> list[int]:
    """
    Для насосов ищем не только точный тег К5, но и под-теги шаблона:
    К5.1, К5.2, К5.3 и т.п.
    Это нужно, потому что Excel-шаблон уже содержит технологические
    обозначения в наименовании строк.
    """
    t = _norm_text(pump_tag)
    m = re.fullmatch(r"К(\d+)", t, flags=re.IGNORECASE)
    if not m:
        return []

    num = m.group(1)
    pat = re.compile(rf"(?<!\w)К{num}(?:\.\d+)?(?!\w)", re.IGNORECASE)

    rows = []
    for r in range(start_row, end_row + 1):
        v_name = ws[f"{COL_NAME}{r}"].value
        v_type = ws[f"{COL_TYPE}{r}"].value

        name = str(v_name or "")
        typ = str(v_type or "")

        if not name and not typ:
            continue
        if _row_is_service(name) or RESERVE_NAME_RE.match(name):
            continue

        if pat.search(name) or pat.search(typ):
            rows.append(r)

    return rows

def _find_row_by_family(ws, family_key: str, *, start_row: int, end_row: int) -> int | None:
    pat = PUMP_FAMILIES.get(family_key)
    if not pat:
        return None
    for r in range(start_row, end_row + 1):
        name = ws[f"{COL_NAME}{r}"].value
        if not name:
            continue
        if _row_is_service(name) or RESERVE_NAME_RE.match(str(name)):
            continue
        if pat.search(str(name)):
            return r
    return None

def _find_row_by_model(ws, model: str, *, start_row: int, end_row: int) -> int | None:
    """
    Универсальный дедуп: ищем существующую строку по модели насоса в тексте наименования (колонка D).
    Работает для любых проектов, если в display_name Excel обычно фигурирует модель.
    """
    mkey = _clean_ws(model).lower()
    if not mkey:
        return None

    # нормализуем: убираем двойные пробелы, приводим к нижнему регистру
    for r in range(start_row, end_row + 1):
        name = ws[f"{COL_NAME}{r}"].value
        if not name:
            continue
        if _row_is_service(name) or RESERVE_NAME_RE.match(str(name)):
            continue
        n = _clean_ws(str(name)).lower()
        if mkey in n:
            return r
    return None

def _get_next_m_index(ws, *, start_row: int, end_row: int) -> int:
    mx = 0
    for r in range(start_row, end_row + 1):
        v = ws[f"{COL_TYPE}{r}"].value
        if not v:
            continue
        m = re.fullmatch(r"M(\d+)", str(v).strip(), re.IGNORECASE)
        if m:
            mx = max(mx, int(m.group(1)))
    return mx + 1


def _infer_type_for_row(display_name: str, ep_kind: str | None, next_m: int) -> tuple[str | None, int]:
    name = (display_name or "").lower().strip()

    if RESERVE_NAME_RE.match(name):
        return "-", next_m

    if ep_kind == "pump" or "насос" in name:
        typ = f"M{next_m}"
        return typ, next_m + 1

    if "шкаф сетей связи" in name:
        return "ШСС", next_m
    if "водонагреватель" in name:
        return "ВН", next_m

    return None, next_m


def ensure_space_and_insert(ws, *, busbar_start: int, log: list[str]) -> int:
    reserve_row = find_reserve_row(ws, start_row=LOADS_START_ROW, end_row=busbar_start - 1)
    anchor = reserve_row or busbar_start

    for r in range(anchor - 1, DATA_START_ROW - 1, -1):
        if ws[f"{COL_NAME}{r}"].value in (None, ""):
            return r

    insert_at = anchor
    ws.insert_rows(insert_at, amount=1)
    _copy_row_style(ws, insert_at - 1, insert_at, min_col=1, max_col=60)
    log.append(f"[INSERT_ROW] inserted row at {insert_at} (before {'RESERVE' if reserve_row else 'BUSBARS'})")
    return insert_at


def _normalize_reserve_row(ws, reserve_row: int, log: list[str]):
    ws[f"{COL_KPD}{reserve_row}"].value = 100
    ws[f"{COL_COS}{reserve_row}"].value = 0.95
    if ws[f"{COL_KI}{reserve_row}"].value in (None, ""):
        ws[f"{COL_KI}{reserve_row}"].value = 1
    apply_load_row_formulas(ws, reserve_row, log)


def _apply_ukrm(ws, *, end_row: int, log: list[str], target_cos: float = 0.95):
    """
    Подбор УКРМ по секциям шин:
    - Находим строку "Секция шин X (максимальный режим ...)" -> base_row
    - Ниже должна быть строка "Регулируемое устройство компенсации реактивной мощности" -> dev_row
    - Ниже должна быть строка "... с учетом УКРМ" -> with_row
    Дальше:
      kvar = CEILING(MAX(0, Q_base*(tg_base - tg_target)), 5)
      ток УКРМ считаем от kvar в строке dev_row:  Iukrm = (R_dev*1000)/(1.73*380)
      строка with_row пересчитывается с вычитанием kvar из реактивной мощности.
    """
    tg_target = math.sqrt(1 - target_cos ** 2) / target_cos
    
    max_dev_row_by_sec: dict[int, int] = {}

    for r in range(1, end_row + 1):
        v = ws[f"{COL_NAME}{r}"].value
        if not v:
            continue
        txt = str(v).lower()

        # базовая секция шин (без УКРМ)
        if "секция шин" in txt and "с учетом укрм" not in txt and "итого" not in txt:
            base_row = r

            # найти строку устройства УКРМ
            dev_row = None
            for rr in range(base_row + 1, min(base_row + 6, end_row + 1)):
                vv = ws[f"{COL_NAME}{rr}"].value
                if vv and ("устройство компенсации реактивной мощности" in str(vv).lower() or "компенсация реактивной мощности" in str(vv).lower()):
                    dev_row = rr
                    break

            # найти строку "секция шин ... с учетом УКРМ"
            with_row = None
            for rr in range(base_row + 1, min(base_row + 10, end_row + 1)):
                vv = ws[f"{COL_NAME}{rr}"].value
                if vv and "с учетом укрм" in str(vv).lower():
                    with_row = rr
                    break

            if dev_row is None or with_row is None:
                continue

            # kvar в строке dev_row
            # определяем режим и секцию (11/12)
            base_text = str(ws[f"{COL_NAME}{base_row}"].value or "").lower()
            sec = _parse_section(ws[f"{COL_SECTION}{base_row}"].value) or 11
            is_normal = "нормальный режим" in base_text
            is_max = "максимальный режим" in base_text

            # "максимальный режим" считаем как раньше и запоминаем dev_row для секции
            if is_max:
                max_dev_row_by_sec[sec] = dev_row

            # "нормальный режим" — используем тот же УКРМ, что и в максимальном режиме этой секции
            if is_normal and sec in max_dev_row_by_sec:
                ref_dev = max_dev_row_by_sec[sec]
                ws[f"{COL_UKRM_KVAR}{dev_row}"].value = f"={COL_UKRM_KVAR}{ref_dev}"
                log.append(f"[UKRM_REF] normal dev_row={dev_row} -> {COL_UKRM_KVAR}{ref_dev} (sec={sec})")
            else:
                # стандартный расчет kvar (как было)
                ws[f"{COL_UKRM_KVAR}{dev_row}"].value = (
                    f"=CEILING(MAX(0,{COL_SEC_P_ACTIVE}{base_row}*(K{base_row}-{tg_target:.6f})),5)"
                )

            # !!! ВАЖНО: ток УКРМ берём от R{dev_row}, а не от base_row
            ws[f"{COL_SEC_I_TOTAL}{dev_row}"].value = f"=({COL_UKRM_KVAR}{dev_row}*1000)/(1.73*380)"

            # строка "с учетом УКРМ"
            ws[f"{COL_SEC_P_ACTIVE}{with_row}"].value = f"={COL_SEC_P_ACTIVE}{base_row}"
            ws[f"{COL_SEC_Q_REACTIVE}{with_row}"].value = f"=MAX(0,{COL_SEC_Q_REACTIVE}{base_row}-{COL_UKRM_KVAR}{dev_row})"
            ws[f"{COL_SEC_S_APPARENT}{with_row}"].value = f"=SQRT({COL_SEC_P_ACTIVE}{with_row}^2+{COL_SEC_Q_REACTIVE}{with_row}^2)"
            ws[f"{COL_SEC_I_TOTAL}{with_row}"].value = f"={COL_SEC_S_APPARENT}{with_row}/(0.38*SQRT(3))"
            ws[f"{COL_COS}{with_row}"].value = f"={COL_SEC_P_ACTIVE}{with_row}/{COL_SEC_S_APPARENT}{with_row}"
            ws[f"{COL_TG}{with_row}"].value = f"=ROUND(SQRT(1-({COL_COS}{with_row})^2)/{COL_COS}{with_row},2)"

            log.append(f"[UKRM] base_row={base_row} dev_row={dev_row} with_row={with_row}")
            
def _find_total_vru_row(ws, *, end_row: int = 5000) -> int | None:
    for r in range(1, end_row + 1):
        v = ws[f"{COL_NAME}{r}"].value
        if v and "итого вру" in str(v).lower():
            return r
    return None

def _find_busbar_rows_with_ukrm(ws, *, end_row: int = 5000) -> dict[int, int]:
    """
    Возвращает {секция: row} для строк вида:
      'Итого секция шин X ... с учетом УКРМ'
    Берем последние найденные строки по секциям.
    """
    out: dict[int, int] = {}

    for r in range(1, end_row + 1):
        v = ws[f"{COL_NAME}{r}"].value
        if not v:
            continue

        txt = str(v).lower()
        if "секция шин" not in txt:
            continue
        if "с учетом укрм" not in txt:
            continue

        sec = _parse_section(ws[f"{COL_SECTION}{r}"].value)
        if sec in (11, 12):
            out[sec] = r

    return out

def _rebuild_total_vru_formulas(ws, *, end_row: int, log: list[str]):
    """
    Восстанавливает формулы строки 'Итого ВРУ' после вставки новых ЭП.

    Логика:
    - Итого ВРУ должно суммировать секции ПОСЛЕ компенсации, то есть
      строки '... с учетом УКРМ'
    - затем:
        S = SQRT(Q^2 + R^2)
        T = S / (0.38 * SQRT(3))
        J = Q / S
        K = ROUND(SQRT(1-J^2)/J, 2)
    """
    r = _find_total_vru_row(ws, end_row=end_row)
    if r is None:
        log.append("[TOTAL_VRU_SKIP] row not found")
        return

    sec_rows = _find_busbar_rows_with_ukrm(ws, end_row=end_row)
    row11 = sec_rows.get(11)
    row12 = sec_rows.get(12)

    if not row11 or not row12:
        log.append(f"[TOTAL_VRU_SKIP] missing with_ukrm rows: sec11={row11}, sec12={row12}")
        return

    old_q = ws[f"{COL_SEC_P_ACTIVE}{r}"].value
    old_r = ws[f"{COL_SEC_Q_REACTIVE}{r}"].value
    old_s = ws[f"{COL_SEC_S_APPARENT}{r}"].value
    old_t = ws[f"{COL_SEC_I_TOTAL}{r}"].value
    old_j = ws[f"{COL_COS}{r}"].value
    old_k = ws[f"{COL_TG}{r}"].value

    # Активная и реактивная мощность ВРУ = сумма секций ПОСЛЕ УКРМ
    ws[f"{COL_SEC_P_ACTIVE}{r}"].value = f"={COL_SEC_P_ACTIVE}{row11}+{COL_SEC_P_ACTIVE}{row12}"
    ws[f"{COL_SEC_Q_REACTIVE}{r}"].value = f"={COL_SEC_Q_REACTIVE}{row11}+{COL_SEC_Q_REACTIVE}{row12}"

    # Полная мощность, ток, cosφ, tgφ
    ws[f"{COL_SEC_S_APPARENT}{r}"].value = f"=SQRT({COL_SEC_P_ACTIVE}{r}^2+{COL_SEC_Q_REACTIVE}{r}^2)"
    ws[f"{COL_SEC_I_TOTAL}{r}"].value = f"={COL_SEC_S_APPARENT}{r}/(0.38*SQRT(3))"
    ws[f"{COL_COS}{r}"].value = f"={COL_SEC_P_ACTIVE}{r}/{COL_SEC_S_APPARENT}{r}"
    ws[f"{COL_TG}{r}"].value = f"=ROUND(SQRT(1-({COL_COS}{r})^2)/{COL_COS}{r},2)"

    log.append(
        f"[TOTAL_VRU_SRC] row={r} sec11_row={row11} sec12_row={row12}"
    )
    log.append(
        f"[TOTAL_VRU] row={r} Q:{old_q}->{ws[f'{COL_SEC_P_ACTIVE}{r}'].value}"
    )
    log.append(
        f"[TOTAL_VRU] row={r} R:{old_r}->{ws[f'{COL_SEC_Q_REACTIVE}{r}'].value}"
    )
    log.append(
        f"[TOTAL_VRU] row={r} S:{old_s}->{ws[f'{COL_SEC_S_APPARENT}{r}'].value}"
    )
    log.append(
        f"[TOTAL_VRU] row={r} T:{old_t}->{ws[f'{COL_SEC_I_TOTAL}{r}'].value}"
    )
    log.append(
        f"[TOTAL_VRU] row={r} J:{old_j}->{ws[f'{COL_COS}{r}'].value}"
    )
    log.append(
        f"[TOTAL_VRU] row={r} K:{old_k}->{ws[f'{COL_TG}{r}'].value}"
    )
# ===== ДОБАВЛЕНО: мягкая очистка (soft delete) лишних строк и дубликатов =====

TAG_IN_PARENS_RE = re.compile(r"\(([^()]+)\)")
# "похожие на теги" обозначения в шаблоне
LIKELY_TAG_RE = re.compile(
    r"^(?:"
    r"(?:К\d{1,3}(?:\.\d{1,2})?)|"     # К4, К7.2
    r"(?:ГГ\.\d{1,3})|"               # ГГ.1
    r"(?:А\d{1,3})|"                  # А1
    r"(?:П\d{1,3})|"                  # П1
    r"(?:В\d{1,3})|"                  # В1
    r"(?:ГР\.\d)|"                    # Гр.1 (вверх)
    r"(?:Гр\.\d)|"                    # Гр.1 (как есть)
    r"(?:EK\d{1,3})|"                 # EK1
    r"(?:Ш[А-ЯA-Z0-9]{2,})|"          # ШУК, ШУТ, ШОВ, ЩНО...
    r"(?:ХВО)|"                       # ХВО
    r"(?:ВН)|"                        # водонагреватель если у тебя так
    r")$",
    re.IGNORECASE
)

def _extract_row_tags(ws, r: int) -> tuple[set[str], set[str]]:
    """
    Пытаемся понять, какие "теги" относятся к строке:
    - если в COL_TYPE что-то похожее на тег (не M*)
    - если в имени есть (...), извлекаем содержимое скобок и берём похожее на тег
    """
    full_tags: set[str] = set()
    canon_tags: set[str] = set()

    v_type = ws[f"{COL_TYPE}{r}"].value
    if v_type:
        s = str(v_type).strip()
        # типы насосов M1/M2 не считаем тегами
        if not re.fullmatch(r"M\d+", s, re.IGNORECASE) and LIKELY_TAG_RE.match(s):
            full_tags.add(_full_tag(s))
            canon_tags.add(_canon_tag_for_keep(s))

    v_name = ws[f"{COL_NAME}{r}"].value
    if v_name:
        name = str(v_name)
        # ищем конструкции "(К7.2)" или "(Гр.1)" и т.п.
        for m in TAG_IN_PARENS_RE.finditer(name):
            cand = m.group(1).strip()
            if LIKELY_TAG_RE.match(cand):
                full_tags.add(_full_tag(cand))
                canon_tags.add(_canon_tag_for_keep(cand))

        # иногда tag пишут прямо в начале/середине строки без скобок
        for m in re.finditer(
            r"\b(К\d{1,3}(?:\.\d{1,2})?|ГГ\.\d{1,3}|А\d{1,3}|П\d{1,3}|В\d{1,3}|Гр\.\d|EK\d{1,3}|ШК\d{1,3}|ЩУТ|ЩУГ|ШУТ|ШУГ|ШУК|ШСС|ХВО)\b",
            name,
            re.IGNORECASE,
        ):
            tg = m.group(1)
            full_tags.add(_full_tag(tg))
            canon_tags.add(_canon_tag_for_keep(tg))

    return full_tags, canon_tags


def _soft_zero_row(ws, r: int, log: list[str], reason: str):
    """
    Мягко "выключаем" строку из расчёта:
    - Ki = 0
    - P = 0
    - имя помечаем, но не трогаем формулы и не удаляем строки
    """
    # Ki
    old_ki = ws[f"{COL_KI}{r}"].value
    ws[f"{COL_KI}{r}"].value = 0

    # P
    old_p = ws[f"{COL_P}{r}"].value
    ws[f"{COL_P}{r}"].value = 0

    # имя
    old_name = ws[f"{COL_NAME}{r}"].value
    nm = _clean_ws(str(old_name or ""))
    if nm and "не используется" not in nm.lower():
        ws[f"{COL_NAME}{r}"].value = f"{nm} (не используется)"
    log.append(f"[PRUNE_ZERO] row={r} reason={reason} P:{old_p}->0 Ki:{old_ki}->0")


def _prune_unused_and_duplicates(ws, *, busbar_start: int, keep_tags: set[str], log: list[str]):
    """
    1) Находим дубликаты строк по тегу -> оставляем первую, остальные зануляем.
    2) Находим строки, которые не относятся ни к одному keep_tag -> зануляем.
    Работает только в блоке нагрузок (LOADS_START_ROW..busbar_start-1).
    """
    start = LOADS_START_ROW
    end = busbar_start - 1

    # карта tag -> список строк
    # карта full_tag -> список строк (только для дедупа)
    tag_rows_full: dict[str, list[int]] = {}

    # cache tags per row
    row_full: dict[int, set[str]] = {}
    row_canon: dict[int, set[str]] = {}

    for r in range(start, end + 1):
        name = ws[f"{COL_NAME}{r}"].value
        if not name:
            continue
        if _row_is_service(str(name)) or RESERVE_NAME_RE.match(str(name)):
            continue

        full_tags, canon_tags = _extract_row_tags(ws, r)
        row_full[r] = full_tags
        row_canon[r] = canon_tags

        for t in full_tags:
            tag_rows_full.setdefault(t, []).append(r)

    # 1) Дубликаты: только по full_tag
    for t, rows in tag_rows_full.items():
        if len(rows) > 1:
            rows_sorted = sorted(rows)
            keep_r = rows_sorted[0]
            for rr in rows_sorted[1:]:
                _soft_zero_row(ws, rr, log, reason=f"duplicate_of_{t}_keep_row_{keep_r}")

    # 2) Лишние: если строка не содержит ни одного canon_tag из keep_tags
    for r in range(start, end + 1):
        name = ws[f"{COL_NAME}{r}"].value
        if not name:
            continue
        if _row_is_service(str(name)) or RESERVE_NAME_RE.match(str(name)):
            continue

        canon_tags = row_canon.get(r) or set()
        if not (canon_tags & keep_tags):
            _soft_zero_row(ws, r, log, reason="not_in_keep_set")

def upsert_items_to_template(
    *,
    template_path: Path,
    out_path: Path,
    items: list[dict],
    registry: Optional[list[dict]] = None,
    overwrite: bool = False,
    only_fill_empty: bool = True,
    end_row: int = 3000,
    audit_csv_path: Optional[Path] = None,
    prune_template: bool = False,
) -> list[str]:
    wb = load_workbook(template_path)
    ws = wb[SHEET_NAME]
    log: list[str] = []

    audit_file = None
    audit_writer = None
    if audit_csv_path:
        audit_file = open(audit_csv_path, "w", newline="", encoding="utf-8")
        audit_writer = csv.writer(audit_file)
        audit_writer.writerow(["tag", "cell", "old", "new", "source"])

    try:
        busbar_start = find_busbar_block_start(ws, end_row=end_row)
        if busbar_start is None:
            raise RuntimeError("Не найден блок 'Секция шин'.")

        next_m = _get_next_m_index(ws, start_row=LOADS_START_ROW, end_row=busbar_start - 1)
        

        # текущие суммы для выбора секции/фазы при вставке
        section_p_sums = {11: 0.0, 12: 0.0}
        phase_p_sums = {11: {"L1": 0.0, "L2": 0.0, "L3": 0.0}, 12: {"L1": 0.0, "L2": 0.0, "L3": 0.0}}
        for r in range(LOADS_START_ROW, busbar_start):
            sec = _parse_section(ws[f"{COL_SECTION}{r}"].value)
            p_val = ws[f"{COL_P}{r}"].value
            try:
                p = float(p_val) if p_val not in (None, "") else None
            except Exception:
                p = None
            if sec in (11, 12) and p is not None:
                section_p_sums[sec] += p
                ph = _parse_phase(ws[f"{COL_PHASES}{r}"].value)
                if ph:
                    phase_p_sums[sec][ph] += p

        for it in items:
            tag = it.get("tag")
            if not tag:
                continue

            display_name = it.get("display_name") or tag
            ep_kind = it.get("ep_kind")
            model = it.get("model")

            u_v = it.get("u_v")
            p_kw = it.get("p_kw")
            kpd = it.get("kpd") or it.get("eta_pct")
            cos = it.get("cos_phi")
            phases = it.get("phases")

            force_overwrite = bool(it.get("_force_overwrite", False))
            eff_overwrite = force_overwrite or (overwrite and (not only_fill_empty))
            
            # "Жёсткая" перезапись базовых параметров для расчётных нагрузок:
            # иначе остаются шаблонные значения (как у тебя случилось с Гр.1/Гр.2 и U=220, P=0.5, cos=1)
            template_group = str(it.get("template_group") or "").strip().lower()

            hard_overwrite_load = (
                ep_kind in ("pump", "hvac", "lighting", "heating", "cabinet", "burner")
                or template_group in ("pump", "hvac_aggregate", "hvac_fan", "exhaust_fan", "lighting", "heating_trace", "boiler_cabinet", "burner", "fixed_tag")
            ) or (it.get("source") == "eo") or (tag == "ХВО")
            
            excel_tag = _excel_tag_from_item(it)

            rows = []

            # 0) для насосов сначала пытаемся найти ВСЕ строки группы по технологическому тегу:
            # К4 -> К4.1, К4.2, ...
            # К5 -> К5.1, К5.2, ...
            if template_group == "pump":
                rows = _find_pump_group_rows(ws, tag, start_row=DATA_START_ROW, end_row=end_row)
                if rows:
                    log.append(f"[MATCH_PUMP_GROUP] {tag} rows={rows}")

            # 1) если не нашли — пробуем excel_tag
            if not rows and excel_tag:
                rows = _find_rows_by_excel_tag(ws, excel_tag, start_row=DATA_START_ROW, end_row=end_row)
                if rows:
                    log.append(f"[MATCH_EXCEL_TAG] {tag} -> '{excel_tag}' rows={rows}")

            # 2) если всё ещё не нашли — старое совпадение по исходному tag
            if not rows:
                rows = _find_rows_by_tag_any(ws, tag, start_row=DATA_START_ROW, end_row=end_row)
            
            # 2) дедуп по семейству (и по model если display_name == "К9")
            # 2) дедуп: сначала по модели (универсально), затем по "семейству" (fallback)
            matched_by_family = False
            matched_by_model = False

            if not rows and ep_kind == "pump" and model:
                r_mod = _find_row_by_model(ws, model, start_row=LOADS_START_ROW, end_row=busbar_start - 1)
                if r_mod is not None:
                    rows = [r_mod]
                    matched_by_model = True
                    log.append(f"[DEDUP_MODEL] {tag}: matched row={r_mod} model='{model}' name='{display_name}'")

            if not rows:
                fam = _family_key_from_item(display_name, model)
                if fam:
                    r_fam = _find_row_by_family(ws, fam, start_row=LOADS_START_ROW, end_row=busbar_start - 1)
                    if r_fam is not None:
                        rows = [r_fam]
                        matched_by_family = True
                        log.append(f"[DEDUP_FAMILY] {tag}: matched row={r_fam} family='{fam}' name='{display_name}' model='{model}'")
            
            if not rows:
                # вставка новой строки
                r = ensure_space_and_insert(ws, busbar_start=busbar_start, log=log)
                busbar_start = find_busbar_block_start(ws, end_row=end_row) or busbar_start

                ws[f"{COL_NAME}{r}"].value = display_name
                typ, next_m = _infer_type_for_row(display_name, ep_kind, next_m)

                # Приоритет у нормализованного обозначения Excel
                if excel_tag:
                    typ = excel_tag
                elif not typ:
                    typ = tag
                    
                if excel_tag:
                    log.append(f"[SET_EXCEL_TAG] {tag} -> {excel_tag}")
                
                if ws[f"{COL_TYPE}{r}"].value in (None, "") and typ:
                    ws[f"{COL_TYPE}{r}"].value = typ

                if excel_tag and ws[f"{COL_TYPE}{r}"].value in (None, ""):
                    ws[f"{COL_TYPE}{r}"].value = excel_tag
                    log.append(f"[FILL_TYPE] {tag} row={r} type='{excel_tag}'")
                
                _set_cell(ws, f"{COL_U}{r}", u_v, overwrite=True, log=log, tag=tag, audit_writer=audit_writer, source="passport_or_user")
                _set_cell(ws, f"{COL_P}{r}", p_kw, overwrite=True, log=log, tag=tag, audit_writer=audit_writer, source="passport_or_user")
                if kpd is not None:
                    _set_cell(ws, f"{COL_KPD}{r}", kpd, overwrite=True, log=log, tag=tag, audit_writer=audit_writer, source="passport_or_user")
                if cos is not None:
                    _set_cell(ws, f"{COL_COS}{r}", cos, overwrite=True, log=log, tag=tag, audit_writer=audit_writer, source="passport_or_user")

                if ws[f"{COL_SECTION}{r}"].value in (None, ""):
                    sec = _choose_section(section_p_sums)
                    ws[f"{COL_SECTION}{r}"].value = sec
                else:
                    sec = _parse_section(ws[f"{COL_SECTION}{r}"].value) or 11

                if phases == 3:
                    ws[f"{COL_PHASES}{r}"].value = "L1, L2, L3"
                elif phases == 1:
                    ws[f"{COL_PHASES}{r}"].value = _choose_phase(phase_p_sums[sec])

                if p_kw not in (None, "", 0):
                    apply_load_row_formulas(ws, r, log)

                log.append(f"[APPEND] {tag} -> row {r}")
                continue

            # обновляем найденные строки
            for r in rows:
                if ws[f"{COL_TYPE}{r}"].value in (None, ""):
                    typ, next_m = _infer_type_for_row(display_name, ep_kind, next_m)
                    if typ:
                        ws[f"{COL_TYPE}{r}"].value = typ

                _set_cell(ws, f"{COL_U}{r}", u_v, overwrite=eff_overwrite or hard_overwrite_load, log=log, tag=tag, audit_writer=audit_writer, source="passport_or_user")
                _set_cell(ws, f"{COL_P}{r}", p_kw, overwrite=eff_overwrite or hard_overwrite_load, log=log, tag=tag, audit_writer=audit_writer, source="passport_or_user")

                hard_overwrite_motor = (
                    hard_overwrite_load
                    or ep_kind == "pump"
                    or template_group == "pump"
                    or ("насос" in (display_name or "").lower())
                )
                if kpd is not None:
                    _set_cell(ws, f"{COL_KPD}{r}", kpd, overwrite=(eff_overwrite or hard_overwrite_motor), log=log, tag=tag, audit_writer=audit_writer, source="passport_or_user")
                if cos is not None:
                    _set_cell(ws, f"{COL_COS}{r}", cos, overwrite=(eff_overwrite or hard_overwrite_motor), log=log, tag=tag, audit_writer=audit_writer, source="passport_or_user")

                cur_name = _clean_ws(str(ws[f"{COL_NAME}{r}"].value or ""))
                if tag not in cur_name:
                    ws[f"{COL_NAME}{r}"].value = f"{cur_name} ({tag})".strip()

                if ws[f"{COL_P}{r}"].value not in (None, "", 0):
                    apply_load_row_formulas(ws, r, log)

                mode = "UPDATE"
                if matched_by_model:
                    mode = "UPDATE_MODEL"
                elif matched_by_family:
                    mode = "UPDATE_FAMILY"
                log.append(f"[{mode}] {tag} -> row {r}")

        # нормализация строки "Резерв"
        busbar_start = find_busbar_block_start(ws, end_row=end_row) or busbar_start
        reserve_row = find_reserve_row(ws, start_row=LOADS_START_ROW, end_row=busbar_start - 1)
        if reserve_row is not None:
            if ws[f"{COL_TYPE}{reserve_row}"].value in (None, ""):
                ws[f"{COL_TYPE}{reserve_row}"].value = "-"
            _normalize_reserve_row(ws, reserve_row, log)
            
        # ===== ДОБАВЛЕНО: мягкая очистка лишних строк и дубликатов =====
        if prune_template:
            keep_tags = {_canon_tag_for_keep(it.get("tag")) for it in items if it.get("tag")}
            # всегда сохраняем системные позиции
            keep_tags.update({_canon_tag_for_keep(x) for x in ("Резерв", "ХВО")})
            _prune_unused_and_duplicates(ws, busbar_start=busbar_start, keep_tags=keep_tags, log=log)

        # пересчет формул секций/фаз
        log.extend(update_single_phase_sumifs(ws, end_row=end_row, loads_start=LOADS_START_ROW))
        log.extend(rebuild_busbar_formulas(ws, end_row=end_row, loads_start=LOADS_START_ROW))

        # ✅ подбор УКРМ (после пересчёта секций)
        _apply_ukrm(ws, end_row=end_row, log=log, target_cos=0.95)

        # ✅ восстановление формул итоговой строки ВРУ
        _rebuild_total_vru_formulas(ws, end_row=end_row, log=log)

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return log

    finally:
        if audit_file:
            audit_file.close()